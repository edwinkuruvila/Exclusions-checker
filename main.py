import requests
import glob
import openpyxl
import time
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.styles import Alignment, PatternFill, Font


def get_data_requests(downloadURL):
    cwd = os.path.dirname(os.path.abspath(__file__))
    
    req = requests.get(downloadURL)
    filename = cwd +"/NatExclusions.txt"

    with open(filename, 'wb') as f:
        for chunk in req.iter_content(chunk_size=50000):
            if chunk:
                f.write(chunk)


def get_data_xpath(downloadURL, _xpath):
    # find current directory
    cwd = os.path.dirname(os.path.abspath(__file__))

    # set download directory
    options = webdriver.ChromeOptions()
    prefs = {"download.default_directory":cwd}
    options.add_argument("--headless=new")
    options.add_experimental_option("prefs",prefs)

    driver = webdriver.Chrome(options=options)
    driver.get(downloadURL)

    # wait until item is loaded(up to  20 seconds), then click 
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, _xpath))).click()

    # pause while file is downloading
    while(not glob.glob(cwd+"/*.txt")):
        time.sleep(.1)

    os.rename(glob.glob(cwd+"/*.txt")[0], cwd+'/TexExclusions.txt') 

    driver.close()

def remove_data():
    cwd = os.path.dirname(os.path.abspath(__file__))
    while(glob.glob("*.txt")):
        os.remove(cwd+'/'+glob.glob("*.txt")[0])

    
def update_upload_data():
    LEIE_URL = "https://oig.hhs.gov/exclusions/downloadables/UPDATED.csv"
    
    HHSC_URL = "https://oig.hhsc.state.tx.us/oigportal2/Exclusions/ctl/DOW/mid/384"
    HHSC_XPATH = '//*[@id="dnn_ctr384_DownloadExclusionsFile_lb_DLoad_ExcFile"]'

    remove_data()
    get_data_xpath(HHSC_URL, HHSC_XPATH)
    get_data_requests(LEIE_URL)


def generate_exclusions_list():
    cwd = os.path.dirname(os.path.abspath(__file__))
    all_data = {}
    with open(cwd+"/NatExclusions.txt", "r", encoding='latin-1') as file:
        for line in file:
            line_ = (line.strip().split('"'))
            new_list = [x for x in line_ if x != ","]
            new_list.pop(0)

            if(new_list):
                address = " | Address: " + new_list[9]+' '+new_list[10]+' '+new_list[11]+' '+new_list[12]
                if(new_list[0] and new_list[0]!=' '):
                    if(new_list[2]!= "" and new_list[2]!=" "):
                        name = "Name: " + new_list[1]+ " "+new_list[2]+" "+new_list[0]
                    else:
                        name = "Name: " + new_list[1]+" "+new_list[0]
                    general = " | General: " + new_list[4]+'('+new_list[5]+")"

                    data = (name+address+general).upper()
                    name_key = (new_list[1]+' '+new_list[0]).upper()
                else:
                    name = "Name: " + new_list[3]
                    if(new_list[5] != '' and new_list[5] != ' '):
                        general = " | General: " + new_list[5]
                    else:
                        general = " | General: " + new_list[4]

                    data = (name+address+general).upper()
                    name_key = (new_list[3]).upper()
                    
                if((name_key) not in all_data):
                    all_data[name_key] = [data]
                else:
                    if(data not in all_data[name_key]):
                        all_data[name_key].append(data)

    with open(cwd+"/TexExclusions.txt", "r", encoding='latin-1') as file:
        next(file)
        for line in file:
            line = line.replace('\t', '')
            new_list = (line.strip().split('"'))
            new_list.pop(0)

            if(len(new_list)>2 and (new_list[18]=='' or new_list[18]==' ')):
                if(new_list[2]!='' and new_list[2]!=' '):
                    if(new_list[6]!= "" and new_list[6]!=" "):
                        name = "Name: " + new_list[4]+ " " + new_list[6] + " " + new_list[2]
                    else:
                        name = "Name: " + new_list[4]+" "+new_list[2]

                    if(new_list[10]!='' and new_list[10]!=" "):
                        licenseNum = " | License Number: " + new_list[10]
                    else:
                        licenseNum = " | License Number: NA "

                    general = " | General: " + new_list[8]

                    data = (name+licenseNum+general).upper()
                    name_key = (new_list[4]+' '+new_list[2]).upper()  

                else:
                    name = "Name: " + new_list[0]
                    general = " | General: " + new_list[8]

                    data = (name+general).upper()
                    name_key = (new_list[0]).upper() 

                if((name_key) not in all_data):
                    all_data[name_key] = [data]
                else:
                    if(data not in all_data[name_key]):
                        all_data[name_key].append(data)   
        
    return all_data


def fill_row(sheet, row_num, fill, font):
    for cell in sheet[row_num]:
        cell.fill = fill
        cell.font = font


def set_row_heights(sheet, min_row=2):
    for row in sheet.iter_rows(min_row=min_row):
        max_length = 0
        max_lines = 1  
        for cell in row:
            if cell.value:
                lines = len(str(cell.value).split('\n'))
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
                if lines > max_lines:
                    max_lines = lines
        
        row_height = max(20, (max_length * 0.7) + (max_lines * 15))  
        max_row_height = 50  
        sheet.row_dimensions[row[0].row].height = min(row_height, max_row_height)


def check_exclusions():
    cwd = os.path.dirname(os.path.abspath(__file__))
    # update_upload_data()
    data_set = generate_exclusions_list()

    light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    black_font = Font(color="000000")
    center_alignment = Alignment(horizontal="center", vertical = "center",wrap_text=True)

    wb = openpyxl.load_workbook(cwd+"/Providers_list.xlsx")
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2):
        if (row[0].value):
            input_name = row[0].value.upper()
            if input_name not in data_set:
                row[2].value = "No Exclusions"
                row[2].alignment = Alignment(wrap_text=True)
                fill_row(sheet, row[0].row, light_green_fill, black_font)
            else:
                row[2].alignment = Alignment(wrap_text=True)
                new_text = "Exclusions may exist\nPossible Candidates:"
                for cand_names in data_set[input_name]:
                    new_text += "\n"+(cand_names)
                
                row[2].value = new_text
                fill_row(sheet, row[0].row, light_red_fill, black_font)

    for cell in sheet['A']:
        cell.alignment = center_alignment
    
    set_row_heights(sheet)

    wb.save(cwd+"/Providers_list.xlsx")


check_exclusions()

